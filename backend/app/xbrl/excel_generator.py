"""Generate a multi-sheet Excel workbook from extracted XBRL JSON data.

Mirrors the working-paper structure an accountant would prepare:
  - Summary (company info + reporting period)
  - Balance Sheet (current + prior year comparative)
  - Profit & Loss
  - Cash Flow
  - Directors / Shareholders / Share Capital
  - Property/Plant/Equipment
  - Auditor
  - Disclosures
  - All Facts (flat dump of every extracted value)

Empty/null fields are written as blank cells so the user can fill them in.
"""
from __future__ import annotations

import io
from dataclasses import dataclass
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ── styling shared across sheets ─────────────────────────────────────────────

HEADER_FILL = PatternFill("solid", fgColor="1E3A8A")           # dark blue
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
SECTION_FILL = PatternFill("solid", fgColor="DBEAFE")          # light blue
SECTION_FONT = Font(name="Calibri", size=10, bold=True, color="1E3A8A")
LABEL_FONT = Font(name="Calibri", size=10, bold=True)
VALUE_FONT = Font(name="Calibri", size=10)
EMPTY_FILL = PatternFill("solid", fgColor="FEF3C7")            # pale yellow for blanks
THIN_BORDER = Border(
    left=Side(style="thin", color="CBD5E1"),
    right=Side(style="thin", color="CBD5E1"),
    top=Side(style="thin", color="CBD5E1"),
    bottom=Side(style="thin", color="CBD5E1"),
)


@dataclass
class ExcelResult:
    success: bool
    xlsx_bytes: bytes = b""
    filename: str = "output.xlsx"
    sheet_count: int = 0
    cell_count: int = 0
    error: str = ""


# ── helpers ──────────────────────────────────────────────────────────────────

def _get(data: dict, path: str, default: Any = None) -> Any:
    cur: Any = data
    for seg in path.split("."):
        if not isinstance(cur, dict) or seg not in cur:
            return default
        cur = cur[seg]
    return cur if cur is not None else default


def _write_kv_row(ws, row: int, label: str, value: Any, label_col: int = 1, value_col: int = 2) -> None:
    """Write a label/value row. Blank value gets a yellow fill (visual cue to fill in)."""
    lbl = ws.cell(row=row, column=label_col, value=label)
    lbl.font = LABEL_FONT
    lbl.alignment = Alignment(vertical="center")
    lbl.border = THIN_BORDER

    val = ws.cell(row=row, column=value_col, value=value if value is not None and value != "" else None)
    val.font = VALUE_FONT
    val.alignment = Alignment(vertical="center", wrap_text=True)
    val.border = THIN_BORDER
    if val.value is None:
        val.fill = EMPTY_FILL


def _write_section_header(ws, row: int, title: str, span: int = 2) -> None:
    ws.cell(row=row, column=1, value=title).font = SECTION_FONT
    for c in range(1, span + 1):
        ws.cell(row=row, column=c).fill = SECTION_FILL
        ws.cell(row=row, column=c).border = THIN_BORDER
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)


def _write_column_headers(ws, row: int, headers: list[str]) -> None:
    for col, hdr in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=hdr)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def _autosize(ws, max_width: int = 60) -> None:
    """Approximate auto-fit by inspecting cell value widths."""
    for col in ws.columns:
        try:
            col_letter = col[0].column_letter
        except AttributeError:
            continue  # merged cell ranges
        max_len = 12
        for cell in col:
            if cell.value is None:
                continue
            length = len(str(cell.value))
            if length > max_len:
                max_len = length
        ws.column_dimensions[col_letter].width = min(max_len + 2, max_width)


# ── sheet writers ────────────────────────────────────────────────────────────

def _sheet_summary(wb: Workbook, data: dict) -> None:
    ws = wb.create_sheet("Summary")
    ws.cell(row=1, column=1, value="Tawthiq XBRL Extraction — Summary").font = Font(size=14, bold=True, color="1E3A8A")
    ws.merge_cells("A1:B1")
    ws.row_dimensions[1].height = 22

    row = 3
    _write_section_header(ws, row, "Company Information"); row += 1
    fields = [
        ("Company Name", _get(data, "company.name")),
        ("CIN", _get(data, "company.cin")),
        ("PAN", _get(data, "company.pan")),
        ("Registered Address", _get(data, "company.registered_address")),
        ("State / UT", _get(data, "company.state")),
        ("Industry Type", _get(data, "company.industry_type")),
        ("Company Type", _get(data, "company.company_type")),
        ("Incorporation Date", _get(data, "company.incorporation_date")),
        ("Email", _get(data, "company.email")),
        ("Telephone", _get(data, "company.telephone")),
        ("Website", _get(data, "company.website")),
    ]
    for label, value in fields:
        _write_kv_row(ws, row, label, value); row += 1

    row += 1
    _write_section_header(ws, row, "Reporting Period"); row += 1
    period_fields = [
        ("Report Type", _get(data, "reporting_period.report_type")),
        ("Period Type", _get(data, "reporting_period.type")),
        ("Current Year — Start", _get(data, "reporting_period.start_date")),
        ("Current Year — End", _get(data, "reporting_period.end_date")),
        ("Prior Year — Start", _get(data, "reporting_period.prior_year_start")),
        ("Prior Year — End", _get(data, "reporting_period.prior_year_end")),
        ("Prior-Prior Year End", _get(data, "reporting_period.prior_prior_year_end")),
        ("Level of Rounding", _get(data, "reporting_period.level_of_rounding")),
        ("Cash Flow Method", _get(data, "reporting_period.cash_flow_method")),
    ]
    for label, value in period_fields:
        _write_kv_row(ws, row, label, value); row += 1

    row += 1
    _write_section_header(ws, row, "Board Approval"); row += 1
    _write_kv_row(ws, row, "Date of Board Meeting", _get(data, "board_approval.date_of_board_meeting")); row += 1
    _write_kv_row(ws, row, "Date of Signing BS", _get(data, "board_approval.date_of_signing_balance_sheet")); row += 1

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 50


def _sheet_balance_sheet(wb: Workbook, data: dict) -> None:
    ws = wb.create_sheet("Balance Sheet")
    cy = _get(data, "reporting_period.end_date", "Current Year")
    py = _get(data, "reporting_period.prior_year_end", "Prior Year")

    ws.cell(row=1, column=1, value="STATEMENT OF FINANCIAL POSITION").font = Font(size=14, bold=True, color="1E3A8A")
    ws.merge_cells("A1:C1")
    ws.row_dimensions[1].height = 22

    _write_column_headers(ws, 3, ["Particulars", str(cy), str(py)])

    rows: list[tuple[str, str, str]] = []   # (label, cy_path, py_path)

    def section(title: str): rows.append((f"§{title}", "", ""))

    # EQUITY & LIABILITIES
    section("I. EQUITY AND LIABILITIES")
    section("  (1) Shareholders' Funds")
    rows.append(("    Share Capital", "balance_sheet.current_year.equity_and_liabilities.share_capital",
                 "balance_sheet.prior_year.equity_and_liabilities.share_capital"))
    rows.append(("    Reserves and Surplus", "balance_sheet.current_year.equity_and_liabilities.reserves_and_surplus",
                 "balance_sheet.prior_year.equity_and_liabilities.reserves_and_surplus"))
    rows.append(("    Money Received Against Share Warrants", "balance_sheet.current_year.equity_and_liabilities.money_received_against_share_warrants",
                 "balance_sheet.prior_year.equity_and_liabilities.money_received_against_share_warrants"))
    section("  (2) Non-Current Liabilities")
    rows.append(("    Long-Term Borrowings", "balance_sheet.current_year.equity_and_liabilities.non_current_liabilities.long_term_borrowings",
                 "balance_sheet.prior_year.equity_and_liabilities.non_current_liabilities.long_term_borrowings"))
    rows.append(("    Deferred Tax Liabilities (Net)", "balance_sheet.current_year.equity_and_liabilities.non_current_liabilities.deferred_tax_liabilities",
                 "balance_sheet.prior_year.equity_and_liabilities.non_current_liabilities.deferred_tax_liabilities"))
    rows.append(("    Other Long-Term Liabilities", "balance_sheet.current_year.equity_and_liabilities.non_current_liabilities.other_long_term_liabilities",
                 "balance_sheet.prior_year.equity_and_liabilities.non_current_liabilities.other_long_term_liabilities"))
    rows.append(("    Long-Term Provisions", "balance_sheet.current_year.equity_and_liabilities.non_current_liabilities.long_term_provisions",
                 "balance_sheet.prior_year.equity_and_liabilities.non_current_liabilities.long_term_provisions"))
    section("  (3) Current Liabilities")
    rows.append(("    Short-Term Borrowings", "balance_sheet.current_year.equity_and_liabilities.current_liabilities.short_term_borrowings",
                 "balance_sheet.prior_year.equity_and_liabilities.current_liabilities.short_term_borrowings"))
    rows.append(("    Trade Payables", "balance_sheet.current_year.equity_and_liabilities.current_liabilities.trade_payables",
                 "balance_sheet.prior_year.equity_and_liabilities.current_liabilities.trade_payables"))
    rows.append(("    Other Current Liabilities", "balance_sheet.current_year.equity_and_liabilities.current_liabilities.other_current_liabilities",
                 "balance_sheet.prior_year.equity_and_liabilities.current_liabilities.other_current_liabilities"))
    rows.append(("    Short-Term Provisions", "balance_sheet.current_year.equity_and_liabilities.current_liabilities.short_term_provisions",
                 "balance_sheet.prior_year.equity_and_liabilities.current_liabilities.short_term_provisions"))
    rows.append(("TOTAL EQUITY AND LIABILITIES", "balance_sheet.current_year.equity_and_liabilities.total_equity_and_liabilities",
                 "balance_sheet.prior_year.equity_and_liabilities.total_equity_and_liabilities"))

    # ASSETS
    section("II. ASSETS")
    section("  (1) Non-Current Assets")
    rows.append(("    Tangible Assets", "balance_sheet.current_year.assets.non_current_assets.tangible_assets",
                 "balance_sheet.prior_year.assets.non_current_assets.tangible_assets"))
    rows.append(("    Intangible Assets", "balance_sheet.current_year.assets.non_current_assets.intangible_assets",
                 "balance_sheet.prior_year.assets.non_current_assets.intangible_assets"))
    rows.append(("    Capital Work-in-Progress", "balance_sheet.current_year.assets.non_current_assets.capital_work_in_progress",
                 "balance_sheet.prior_year.assets.non_current_assets.capital_work_in_progress"))
    rows.append(("    Non-Current Investments", "balance_sheet.current_year.assets.non_current_assets.non_current_investments",
                 "balance_sheet.prior_year.assets.non_current_assets.non_current_investments"))
    rows.append(("    Deferred Tax Assets (Net)", "balance_sheet.current_year.assets.non_current_assets.deferred_tax_assets",
                 "balance_sheet.prior_year.assets.non_current_assets.deferred_tax_assets"))
    rows.append(("    Long-Term Loans and Advances", "balance_sheet.current_year.assets.non_current_assets.long_term_loans_and_advances",
                 "balance_sheet.prior_year.assets.non_current_assets.long_term_loans_and_advances"))
    rows.append(("    Other Non-Current Assets", "balance_sheet.current_year.assets.non_current_assets.other_non_current_assets",
                 "balance_sheet.prior_year.assets.non_current_assets.other_non_current_assets"))
    section("  (2) Current Assets")
    rows.append(("    Current Investments", "balance_sheet.current_year.assets.current_assets.current_investments",
                 "balance_sheet.prior_year.assets.current_assets.current_investments"))
    rows.append(("    Inventories", "balance_sheet.current_year.assets.current_assets.inventories",
                 "balance_sheet.prior_year.assets.current_assets.inventories"))
    rows.append(("    Trade Receivables", "balance_sheet.current_year.assets.current_assets.trade_receivables",
                 "balance_sheet.prior_year.assets.current_assets.trade_receivables"))
    rows.append(("    Cash and Bank Balances", "balance_sheet.current_year.assets.current_assets.cash_and_bank_balances",
                 "balance_sheet.prior_year.assets.current_assets.cash_and_bank_balances"))
    rows.append(("    Short-Term Loans and Advances", "balance_sheet.current_year.assets.current_assets.short_term_loans_and_advances",
                 "balance_sheet.prior_year.assets.current_assets.short_term_loans_and_advances"))
    rows.append(("    Other Current Assets", "balance_sheet.current_year.assets.current_assets.other_current_assets",
                 "balance_sheet.prior_year.assets.current_assets.other_current_assets"))
    rows.append(("TOTAL ASSETS", "balance_sheet.current_year.assets.total_assets",
                 "balance_sheet.prior_year.assets.total_assets"))

    row = 4
    for label, cy_path, py_path in rows:
        if label.startswith("§"):
            _write_section_header(ws, row, label[1:], span=3)
        else:
            is_total = label.strip().startswith("TOTAL")
            _write_kv_row(ws, row, label, _get(data, cy_path), value_col=2)
            _write_kv_row(ws, row, label, _get(data, py_path), label_col=3, value_col=3) if False else None
            # Actually write py to col 3 directly (avoid double-writing label)
            py_cell = ws.cell(row=row, column=3, value=_get(data, py_path))
            py_cell.font = VALUE_FONT
            py_cell.alignment = Alignment(horizontal="right", vertical="center")
            py_cell.border = THIN_BORDER
            if py_cell.value is None:
                py_cell.fill = EMPTY_FILL
            cy_cell = ws.cell(row=row, column=2)
            cy_cell.alignment = Alignment(horizontal="right", vertical="center")
            if is_total:
                for c in (1, 2, 3):
                    ws.cell(row=row, column=c).font = Font(bold=True)
                    ws.cell(row=row, column=c).fill = PatternFill("solid", fgColor="F1F5F9")
        row += 1

    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20


def _sheet_pnl(wb: Workbook, data: dict) -> None:
    ws = wb.create_sheet("Profit & Loss")
    cy = _get(data, "reporting_period.end_date", "Current Year")
    py = _get(data, "reporting_period.prior_year_end", "Prior Year")

    ws.cell(row=1, column=1, value="STATEMENT OF PROFIT OR LOSS").font = Font(size=14, bold=True, color="1E3A8A")
    ws.merge_cells("A1:C1")

    _write_column_headers(ws, 3, ["Particulars", str(cy), str(py)])

    items = [
        ("I. Revenue from Operations", "profit_loss.current_year.revenue_from_operations", "profit_loss.prior_year.revenue_from_operations"),
        ("II. Other Income", "profit_loss.current_year.other_income", "profit_loss.prior_year.other_income"),
        ("III. Total Revenue (I + II)", "profit_loss.current_year.total_revenue", "profit_loss.prior_year.total_revenue"),
        ("§IV. Expenses", "", ""),
        ("    Cost of Materials Consumed", "profit_loss.current_year.expenses.cost_of_materials_consumed", "profit_loss.prior_year.expenses.cost_of_materials_consumed"),
        ("    Purchase of Stock-in-Trade", "profit_loss.current_year.expenses.purchases_of_stock_in_trade", "profit_loss.prior_year.expenses.purchases_of_stock_in_trade"),
        ("    Changes in Inventories", "profit_loss.current_year.expenses.changes_in_inventories", "profit_loss.prior_year.expenses.changes_in_inventories"),
        ("    Employee Benefits Expense", "profit_loss.current_year.expenses.employee_benefits_expense", "profit_loss.prior_year.expenses.employee_benefits_expense"),
        ("    Finance Costs", "profit_loss.current_year.expenses.finance_costs", "profit_loss.prior_year.expenses.finance_costs"),
        ("    Depreciation and Amortisation", "profit_loss.current_year.expenses.depreciation_amortisation", "profit_loss.prior_year.expenses.depreciation_amortisation"),
        ("    Other Expenses", "profit_loss.current_year.expenses.other_expenses", "profit_loss.prior_year.expenses.other_expenses"),
        ("    Total Expenses", "profit_loss.current_year.expenses.total_expenses", "profit_loss.prior_year.expenses.total_expenses"),
        ("V. Profit Before Tax", "profit_loss.current_year.profit_before_tax", "profit_loss.prior_year.profit_before_tax"),
        ("§VI. Tax Expense", "", ""),
        ("    Current Tax", "profit_loss.current_year.tax_expense.current_tax", "profit_loss.prior_year.tax_expense.current_tax"),
        ("    Deferred Tax", "profit_loss.current_year.tax_expense.deferred_tax", "profit_loss.prior_year.tax_expense.deferred_tax"),
        ("    Total Tax Expense", "profit_loss.current_year.tax_expense.total_tax_expense", "profit_loss.prior_year.tax_expense.total_tax_expense"),
        ("VII. Profit / (Loss) for the Period", "profit_loss.current_year.profit_for_period", "profit_loss.prior_year.profit_for_period"),
        ("§VIII. Earnings Per Share", "", ""),
        ("    Basic EPS (Rs.)", "profit_loss.current_year.earnings_per_share.basic", "profit_loss.prior_year.earnings_per_share.basic"),
        ("    Diluted EPS (Rs.)", "profit_loss.current_year.earnings_per_share.diluted", "profit_loss.prior_year.earnings_per_share.diluted"),
    ]

    row = 4
    for label, cy_path, py_path in items:
        if label.startswith("§"):
            _write_section_header(ws, row, label[1:], span=3)
        else:
            is_total = "Profit" in label or "Total" in label
            ws.cell(row=row, column=1, value=label).font = (Font(bold=True) if is_total else LABEL_FONT)
            cy_val = _get(data, cy_path) if cy_path else None
            py_val = _get(data, py_path) if py_path else None
            for col, v in ((2, cy_val), (3, py_val)):
                c = ws.cell(row=row, column=col, value=v)
                c.font = Font(bold=True) if is_total else VALUE_FONT
                c.alignment = Alignment(horizontal="right")
                c.border = THIN_BORDER
                if v is None:
                    c.fill = EMPTY_FILL
            ws.cell(row=row, column=1).border = THIN_BORDER
            if is_total:
                for c in (1, 2, 3):
                    ws.cell(row=row, column=c).fill = PatternFill("solid", fgColor="F1F5F9")
        row += 1

    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 22


def _sheet_cash_flow(wb: Workbook, data: dict) -> None:
    ws = wb.create_sheet("Cash Flow")
    ws.cell(row=1, column=1, value="STATEMENT OF CASH FLOWS").font = Font(size=14, bold=True, color="1E3A8A")
    ws.merge_cells("A1:B1")

    _write_column_headers(ws, 3, ["Particulars", str(_get(data, "reporting_period.end_date", "Current Year"))])

    items = [
        ("A. Cash Flows from Operating Activities", "cash_flow.current_year.cash_from_operating_activities"),
        ("B. Cash Flows from / (used in) Investing Activities", "cash_flow.current_year.cash_from_investing_activities"),
        ("C. Cash Flows from / (used in) Financing Activities", "cash_flow.current_year.cash_from_financing_activities"),
        ("Net Increase / (Decrease) in Cash and Cash Equivalents (A+B+C)", "cash_flow.current_year.net_increase_decrease_in_cash"),
        ("Cash and Cash Equivalents at the Beginning of the Period", "cash_flow.current_year.cash_at_beginning_of_period"),
        ("Cash and Cash Equivalents at the End of the Period", "cash_flow.current_year.cash_at_end_of_period"),
    ]
    row = 4
    for label, path in items:
        ws.cell(row=row, column=1, value=label).font = LABEL_FONT
        ws.cell(row=row, column=1).border = THIN_BORDER
        v = _get(data, path)
        c = ws.cell(row=row, column=2, value=v)
        c.font = VALUE_FONT
        c.alignment = Alignment(horizontal="right")
        c.border = THIN_BORDER
        if v is None:
            c.fill = EMPTY_FILL
        row += 1
    ws.column_dimensions["A"].width = 55
    ws.column_dimensions["B"].width = 22


def _sheet_directors(wb: Workbook, data: dict) -> None:
    ws = wb.create_sheet("Directors")
    ws.cell(row=1, column=1, value="DIRECTORS / KEY MANAGERIAL PERSONNEL").font = Font(size=14, bold=True, color="1E3A8A")
    ws.merge_cells("A1:E1")

    _write_column_headers(ws, 3, ["Name", "DIN", "Designation", "Signs Financial Statements", "Signs Board Report"])
    directors = data.get("directors") or []
    if not directors:
        ws.cell(row=4, column=1, value="(no directors extracted)").font = Font(italic=True, color="64748B")
    for idx, d in enumerate(directors, start=4):
        for col, key in enumerate(["name", "din", "designation", "signs_financial_statements", "signs_board_report"], start=1):
            v = d.get(key) if isinstance(d, dict) else None
            c = ws.cell(row=idx, column=col, value=v)
            c.font = VALUE_FONT
            c.border = THIN_BORDER
            if v is None:
                c.fill = EMPTY_FILL
    _autosize(ws)


def _sheet_shareholders(wb: Workbook, data: dict) -> None:
    ws = wb.create_sheet("Shareholders")
    ws.cell(row=1, column=1, value="SHAREHOLDERS (≥ 5% holding)").font = Font(size=14, bold=True, color="1E3A8A")
    ws.merge_cells("A1:F1")

    _write_column_headers(ws, 3, ["Name", "PAN", "Shares Held", "% Held", "Type of Share", "Country"])
    shareholders = data.get("shareholders") or []
    if not shareholders:
        ws.cell(row=4, column=1, value="(no shareholders extracted)").font = Font(italic=True, color="64748B")
    for idx, s in enumerate(shareholders, start=4):
        if not isinstance(s, dict):
            continue
        for col, key in enumerate(["name", "pan", "shares_held", "percentage_held", "type_of_share", "country_of_residence"], start=1):
            v = s.get(key)
            c = ws.cell(row=idx, column=col, value=v)
            c.font = VALUE_FONT
            c.border = THIN_BORDER
            if v is None:
                c.fill = EMPTY_FILL
    _autosize(ws)


def _sheet_ppe(wb: Workbook, data: dict) -> None:
    ws = wb.create_sheet("PPE Schedule")
    ws.cell(row=1, column=1, value="PROPERTY, PLANT & EQUIPMENT MOVEMENT").font = Font(size=14, bold=True, color="1E3A8A")
    ws.merge_cells("A1:H1")

    headers = [
        "Asset Class", "Opening Gross Block (PY)", "Additions (CY)", "Disposals (CY)",
        "Closing Gross Block (CY)", "Accumulated Depreciation", "Net Carrying Amount (CY)",
        "Useful Life (years)",
    ]
    _write_column_headers(ws, 3, headers)
    items = data.get("property_plant_equipment") or []
    if not items:
        ws.cell(row=4, column=1, value="(no PPE schedule extracted)").font = Font(italic=True, color="64748B")
    for idx, item in enumerate(items, start=4):
        if not isinstance(item, dict):
            continue
        values = [
            item.get("asset_class"),
            item.get("opening_gross_block_current_year") or item.get("opening_gross_block_prior_year"),
            item.get("additions_during_current_year"),
            item.get("disposals_during_current_year"),
            item.get("closing_gross_block_current_year"),
            item.get("accumulated_depreciation_closing"),
            item.get("net_carrying_amount_current_year"),
            item.get("useful_life_years"),
        ]
        for col, v in enumerate(values, start=1):
            c = ws.cell(row=idx, column=col, value=v)
            c.font = VALUE_FONT
            c.border = THIN_BORDER
            if v is None:
                c.fill = EMPTY_FILL
    _autosize(ws)


def _sheet_auditor(wb: Workbook, data: dict) -> None:
    ws = wb.create_sheet("Auditor")
    ws.cell(row=1, column=1, value="AUDITOR INFORMATION").font = Font(size=14, bold=True, color="1E3A8A")
    ws.merge_cells("A1:B1")
    fields = [
        ("Firm Name", _get(data, "auditor.firm_name")),
        ("Firm Registration Number (FRN)", _get(data, "auditor.firm_registration_number")),
        ("Partner Name", _get(data, "auditor.partner_name")),
        ("Membership Number", _get(data, "auditor.membership_number")),
        ("Address", _get(data, "auditor.address")),
        ("Signature Date", _get(data, "auditor.signature_date")),
        ("Opinion Type", _get(data, "auditor.opinion_type")),
    ]
    for r, (label, value) in enumerate(fields, start=3):
        _write_kv_row(ws, r, label, value)
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 60


def _sheet_disclosures(wb: Workbook, data: dict) -> None:
    ws = wb.create_sheet("Disclosures")
    ws.cell(row=1, column=1, value="TEXT DISCLOSURES").font = Font(size=14, bold=True, color="1E3A8A")
    ws.merge_cells("A1:B1")
    disclosures = data.get("disclosures") or {}
    keys = [
        ("Significant Accounting Policies", "significant_accounting_policies"),
        ("Basis of Consolidation", "basis_of_consolidation"),
        ("Employee Benefits", "employee_benefits"),
        ("Related Party Transactions", "related_party_transactions"),
        ("Contingent Liabilities", "contingent_liabilities"),
        ("CSR Activities", "csr_activities"),
    ]
    row = 3
    for label, key in keys:
        _write_kv_row(ws, row, label, disclosures.get(key) if isinstance(disclosures, dict) else None)
        row += 1
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 90


def _sheet_all_facts(wb: Workbook, data: dict) -> None:
    """Flatten every leaf value in the JSON tree → one row per data point."""
    ws = wb.create_sheet("All Facts (flat)")
    ws.cell(row=1, column=1, value="ALL EXTRACTED VALUES (flat dump for debugging / comparison)").font = Font(size=14, bold=True, color="1E3A8A")
    ws.merge_cells("A1:B1")
    _write_column_headers(ws, 3, ["JSON Path", "Value"])

    rows: list[tuple[str, Any]] = []

    def walk(node: Any, prefix: str = "") -> None:
        if isinstance(node, dict):
            for k, v in node.items():
                walk(v, f"{prefix}.{k}" if prefix else k)
        elif isinstance(node, list):
            for i, v in enumerate(node):
                walk(v, f"{prefix}[{i}]")
        else:
            rows.append((prefix, node))

    walk(data)

    for idx, (path, value) in enumerate(rows, start=4):
        c1 = ws.cell(row=idx, column=1, value=path)
        c1.font = Font(name="Consolas", size=9)
        c1.border = THIN_BORDER
        c2 = ws.cell(row=idx, column=2, value=value if value is not None else None)
        c2.font = VALUE_FONT
        c2.border = THIN_BORDER
        c2.alignment = Alignment(wrap_text=True, vertical="top")
        if value is None:
            c2.fill = EMPTY_FILL
    ws.column_dimensions["A"].width = 65
    ws.column_dimensions["B"].width = 60


# ── public API ───────────────────────────────────────────────────────────────

def generate_excel(data: dict, filename_base: str = "tawthiq_xbrl") -> ExcelResult:
    """Build the multi-sheet workbook and return its bytes."""
    try:
        wb = Workbook()
        # remove the default first sheet
        wb.remove(wb.active)

        _sheet_summary(wb, data)
        _sheet_balance_sheet(wb, data)
        _sheet_pnl(wb, data)
        _sheet_cash_flow(wb, data)
        _sheet_directors(wb, data)
        _sheet_shareholders(wb, data)
        _sheet_ppe(wb, data)
        _sheet_auditor(wb, data)
        _sheet_disclosures(wb, data)
        _sheet_all_facts(wb, data)

        buf = io.BytesIO()
        wb.save(buf)
        xlsx_bytes = buf.getvalue()

        # Filename: COMPANY_FY.xlsx
        co = (_get(data, "company.name") or "company").upper()
        co = "".join(c if c.isalnum() else "_" for c in co).strip("_")[:60]
        end = _get(data, "reporting_period.end_date") or ""
        start = _get(data, "reporting_period.start_date") or ""
        fy = ""
        try:
            fy = f"{start[:4]}_{end[2:4]}"
        except Exception:
            fy = (end or "FY").replace("-", "_")
        filename = f"{filename_base}_{co}_{fy}.xlsx"

        # Count cells written
        cell_count = sum(ws.max_row * ws.max_column for ws in wb.worksheets)

        return ExcelResult(
            success=True,
            xlsx_bytes=xlsx_bytes,
            filename=filename,
            sheet_count=len(wb.worksheets),
            cell_count=cell_count,
        )
    except Exception as exc:
        return ExcelResult(success=False, error=str(exc))
