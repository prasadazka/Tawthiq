"""Convert Excel workbook sheets into markdown tables for Gemini consumption."""
from __future__ import annotations

import io
import logging
from typing import Any

from openpyxl import load_workbook

logger = logging.getLogger(__name__)

MAX_ROWS_PER_SHEET = 500
MAX_COLS_PER_SHEET = 20


def _format_cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        if value == int(value):
            return str(int(value))
        return f"{value:.2f}"
    s = str(value).strip()
    return s.replace("|", "\\|").replace("\n", " ")


def excel_to_markdown(xlsx_bytes: bytes) -> dict[str, str]:
    """Load workbook from bytes; return {sheet_name: markdown_table}.

    Empty sheets are skipped. Large sheets are truncated to MAX_ROWS_PER_SHEET.
    """
    wb = load_workbook(io.BytesIO(xlsx_bytes), data_only=True, read_only=True)
    out: dict[str, str] = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        max_row = min(ws.max_row or 0, MAX_ROWS_PER_SHEET)
        max_col = min(ws.max_column or 0, MAX_COLS_PER_SHEET)
        if max_row == 0 or max_col == 0:
            continue

        rows: list[list[str]] = []
        for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col, values_only=True):
            cells = [_format_cell(v) for v in row]
            if any(c for c in cells):
                rows.append(cells)

        if not rows:
            continue

        header = "| " + " | ".join(f"C{i+1}" for i in range(max_col)) + " |"
        sep = "| " + " | ".join(["---"] * max_col) + " |"
        body = "\n".join("| " + " | ".join(r + [""] * (max_col - len(r))) + " |" for r in rows)
        out[sheet_name] = f"{header}\n{sep}\n{body}"

    wb.close()
    return out


def is_valid_excel(xlsx_bytes: bytes) -> bool:
    """Quick magic-byte check: xlsx is a zip (PK\\x03\\x04), xls is OLE2."""
    if len(xlsx_bytes) < 8:
        return False
    return xlsx_bytes[:4] == b"PK\x03\x04" or xlsx_bytes[:8] == b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"
